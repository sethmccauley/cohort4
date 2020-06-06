from openpyxl import load_workbook, Workbook
import os
from collections import defaultdict

localFile = os.getcwd() + '/store_db.xlsx'

def breakIntoDict(file):
    wb = load_workbook(localFile)
    builtDict = {}
    for sheet in wb.sheetnames:
        builtDict[sheet] = {}
        rowCount = 1
        for row in wb[sheet].iter_rows(min_row=2,min_col=1):
            if row[0].value is not None:
                col = 0
                builtDict[sheet][rowCount] = {}
                for cell in row:
                    builtDict[sheet][rowCount][wb[sheet][1][col].value] = cell.value
                    col += 1
            rowCount += 1
    return builtDict

# def innerJoin(dict_one, dict_two, column, value=None):
#     resultingDict = {}
#     index = 1
#     for dOne_rows in dict_one.values():
#         records = 1
#         tempHold = {}
#         for dTwo_rows in dict_two.values():
#             if dOne_rows[column] == dTwo_rows[column]:
#                 tempHold[records] = dTwo_rows
#                 records += 1
#         combined = {}
#         combined.update(dOne_rows)
#         combined.update(tempHold)
#         resultingDict[index] = combined
#         index += 1
#     return resultingDict

def innerJoin(dict_one, dict_two, column, lookup_value=None):
    resultingDict = {}
    index = 1
    # Construct smaller tables to iterate through later if passed a lookup value
    if lookup_value != None:
        dict_one = {key:value for key,value in dict_one.items() if value[column] == lookup_value}
        dict_two = {key:value for key,value in dict_two.items() if value[column] == lookup_value}

    for dOne_rows in dict_one.values():
        for dTwo_rows in dict_two.values():
            if dOne_rows[column] == dTwo_rows[column]:
                combined = {}
                combined.update(dOne_rows)
                combined.update(dTwo_rows)
                resultingDict[index] = combined
                index += 1

    return resultingDict
